import os

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock

import txtparse

BOLD = Font(bold=True)
RED = Font(color="ff0000")

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")

THIN_CASE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM_CASE = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

CENTERED = Alignment(horizontal="center", vertical="center", wrap_text=True)
V_CENTERED = Alignment(horizontal="left", vertical="center", wrap_text=True)

H1_STYLE = NamedStyle(name="h1")
H1_STYLE.font = BOLD
H1_STYLE.fill = PatternFill(fill_type="solid", start_color="8DB3E2", end_color="8DB3E2")
H1_STYLE.alignment = CENTERED
H1_STYLE.border = MEDIUM_CASE

H2_STYLE = NamedStyle(name="h2")
H2_STYLE.font = BOLD
H2_STYLE.fill = PatternFill(fill_type="solid", start_color="C6D9F0", end_color="C6D9F0")
H2_STYLE.border = THIN_CASE


def create_casebook(cv: txtparse.CV) -> Workbook:
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 5
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 30
    ws.column_dimensions["L"].width = 10
    # spacers
    for col in ["A", "D", "I", "M"]:
        ws.column_dimensions[col].width = 2

    ws.merge_cells("E1:J1")
    write(ws, addr="E1", value="签约时间：", font=BOLD, alignment=CENTERED)
    encase(ws, "E1:J1")

    ws.merge_cells("E2:J2")
    write(
        ws,
        addr="E2",
        value="红字项目和奖项在择由规划建议、支持帮助下完成和获得",
        font=RED,
        alignment=CENTERED,
    )
    encase(ws, "E2:J2")

    # name, school, target major
    new_row = create_profile_block(ws, row=4, cv=cv)
    # grades & tests
    new_row = create_grades_block(ws, row=new_row + 1, cv=cv)
    # schools applied to
    create_appications_block(ws, row=new_row + 1)

    # activities
    new_row = create_activities_block(ws, row=4, cv=cv)
    # awards
    create_awards_block(ws, row=new_row + 1, cv=cv)

    # personal brand
    create_h1(ws, "J4", "L4", "个人品牌")
    ws.merge_cells("J5:L7")
    encase(ws, "J5:L7")

    return wb


def create_profile_block(ws, row, cv: txtparse.CV) -> int:
    create_h1(ws, f"B{row}", f"C{row}", "基本信息")
    write(ws, row=row + 1, column="B", value="学生姓名")
    write(ws, row=row + 1, column="C", value=cv.name)
    write(ws, row=row + 2, column="B", value="在读学校")
    write(ws, row=row + 2, column="C", value=cv.last_education.school)
    write(ws, row=row + 3, column="B", value="申请专业")
    write(ws, row=row + 3, column="C")
    encase(ws, f"B{row+1}:C{row+3}")
    # return the next row
    return row + 4


def create_grades_block(ws, row, cv: txtparse.CV) -> int:
    create_h1(ws, f"B{row}", f"C{row}", "硬件")

    write(ws, row=row + 1, column="B", value="年级排名")
    write(ws, row=row + 1, column="C", value=cv.last_education.rank)
    write(ws, row=row + 2, column="B", value="G9成绩")
    write(ws, row=row + 2, column="C")
    write(ws, row=row + 3, column="B", value="G10成绩")
    write(ws, row=row + 3, column="C")
    write(ws, row=row + 4, column="B", value="G11成绩")
    write(ws, row=row + 4, column="C")

    write(ws, row=row + 5, column="B", value="AP/IB", alignment=V_CENTERED)
    courses = cv.last_education.ap_courses() or cv.last_education.ib_courses()
    write(ws, row=row + 5, column="C", value="\n".join(courses), alignment=V_CENTERED)
    ws.merge_cells(f"B{row+5}:B{row+11}")
    ws.merge_cells(f"C{row+5}:C{row+11}")

    try:
        e_test = cv.english_tests()[0]
        e_test_key = e_test.name
        e_test_value = f"{e_test.score} ({e_test.date})"
    except IndexError:
        e_test_key = "TOEFL/IELTS"
        e_test_value = ""
    write(ws, row=row + 12, column="B", value=e_test_key)
    write(ws, row=row + 12, column="C", value=e_test_value)

    try:
        a_test = cv.academic_tests()[0]
        a_test_key = a_test.name
        a_test_value = f"{a_test.score} ({a_test.date})"
    except IndexError:
        a_test_key = "SAT/ACT"
        a_test_value = ""
    write(ws, row=row + 13, column="B", value=a_test_key)
    write(ws, row=row + 13, column="C", value=a_test_value)

    write(ws, row=row + 14, column="B", value="12年级课程", alignment=V_CENTERED)
    write(ws, row=row + 14, column="C", alignment=V_CENTERED)
    ws.merge_cells(f"B{row+14}:B{row+20}")
    ws.merge_cells(f"C{row+14}:C{row+20}")

    encase(ws, f"B{row}:C{row+20}")
    # return the next row
    return row + 21


def create_appications_block(ws, row):
    red_bold = InlineFont(color="ff0000", b=True)
    create_h1(
        ws,
        f"B{row}",
        f"C{row}",
        CellRichText("申请大学", TextBlock(red_bold, "（录取）")),
    )
    write(ws, row=row + 1, column="B", value="ED/REA")
    write(ws, row=row + 1, column="C")
    write(ws, row=row + 2, column="B", value="EA")
    write(ws, row=row + 2, column="C")
    write(ws, row=row + 3, column="B", value="RD", alignment=V_CENTERED)
    ws.merge_cells(f"B{row+3}:B{row+5}")
    ws.merge_cells(f"C{row+3}:C{row+5}")
    encase(ws, f"B{row}:C{row+5}")


def create_activities_block(ws, row, cv: txtparse.CV) -> int:
    create_h1(ws, f"E{row}", f"H{row}", "活动")

    sections = cv.activity_sections
    sections.append("")  # capture unsectioned activities

    r = row + 1
    for section in sections:
        activities = cv.activities_of_section(section)
        if not activities:
            continue
        write(ws, row=r, column="E", value=section, style=H2_STYLE)
        write(ws, row=r, column="F", value="活动描述", style=H2_STYLE)
        write(ws, row=r, column="G", value="时间投入", style=H2_STYLE)
        write(ws, row=r, column="H", value="选择", style=H2_STYLE)
        encase(ws, f"E{r}:H{r}")
        r += 1
        for act in activities:
            write(
                ws,
                row=r,
                column="E",
                value=f"{act.org + ' | ' if act.org else ''}{act.role}",
                alignment=V_CENTERED,
            )
            write(
                ws,
                row=r,
                column="F",
                value=" ".join(act.descriptions),
                alignment=V_CENTERED,
            )
            write(
                ws,
                row=r,
                column="G",
                value=f"{_format_dates(act)}\n{_format_commitment(act)}",
                alignment=V_CENTERED,
            )
            write(ws, row=r, column="H", font=BOLD)
            for col in ["E", "F", "G", "H"]:
                ws.merge_cells(f"{col}{r}:{col}{r+2}")
            encase(ws, f"E{r}:H{r+2}")
            r += 3

    # return the next row
    return r + 1


def create_awards_block(ws, row, cv: txtparse.CV):
    create_h1(ws, f"E{row}", f"H{row}", "奖项")
    write(ws, row=row + 1, column="E", value="年级/层次", style=H2_STYLE)
    write(ws, row=row + 1, column="F", value="奖项名称", style=H2_STYLE)
    ws.merge_cells(f"F{row+1}:G{row+1}")
    write(ws, row=row + 1, column="H", value="选择", style=H2_STYLE)

    r = row + 2
    awards = list(cv.awards)

    while len(awards) < 5:  # ensure minimum 5 rows
        awards.append(txtparse.Award(name=""))

    for award in awards:
        write(ws, row=r, column="E", value=str(award.date))
        write(ws, row=r, column="F", value=award.name)
        ws.merge_cells(f"F{r}:G{r}")
        write(ws, row=r, column="H", font=BOLD)
        r += 1

    # encase from the top of subheading to the bottom of last row
    encase(ws, f"E{row+1}:H{r-1}")


def write(ws, addr="", row="", column="", value="", **kwargs):
    addr = addr or f"{column}{row}"
    if not addr:
        raise ValueError("Empty address")
    ws[addr] = value
    ws[addr].border = THIN_CASE  # thin border by default
    ws[addr].alignment = Alignment(wrap_text=True)  # wrap text by default
    for k, v in kwargs.items():
        setattr(ws[addr], k, v)


def area(ws, *, start="", end=""):
    for row in ws[start:end]:
        yield from row


def create_h1(ws, start, end, value):
    ws.merge_cells(f"{start}:{end}")
    ws[start] = value
    for cell in area(ws, start=start, end=end):
        cell.style = H1_STYLE


def encase(ws, start_end="", *, start="", end=""):
    """Surround the area defined by start:end with medium borders."""
    if start_end:
        start, end = start_end.split(":")
    start_cell = ws[start]
    end_cell = ws[end]

    for cell in area(ws, start=start, end=end):
        border_params = dict(top=THIN, bottom=THIN, left=THIN, right=THIN)
        if cell.row == start_cell.row:
            border_params["top"] = MEDIUM
        if cell.row == end_cell.row:
            border_params["bottom"] = MEDIUM
        if cell.column == start_cell.column:
            border_params["left"] = MEDIUM
        if cell.column == end_cell.column:
            border_params["right"] = MEDIUM
        cell.border = Border(**border_params)


def _format_dates(act: txtparse.Activity) -> str:
    if not act.start_date:
        return str(act.end_date)
    if not act.end_date or (act.start_date == act.end_date):
        return str(act.start_date)
    return f"{act.start_date} - {act.end_date}"


def _format_commitment(act: txtparse.Activity):
    if act.hours_per_week == "1":
        hpw = "1 hr/wk"
    else:
        hpw = f"{act.hours_per_week or '?'} hrs/wk"
    if act.weeks_per_year == "1":
        wpy = "1 wk/yr"
    else:
        wpy = f"{act.weeks_per_year or '?'} wks/yr"
    return f"{hpw}\n{wpy}"


if __name__ == "__main__":
    with open("tests/test_case.txt", encoding="utf-8") as f:
        src = f.read()
    test_cv, _ = txtparse.parse(src)
    wb = create_casebook(test_cv)
    wb.save(os.path.join("tests", "test_case.xlsx"))
